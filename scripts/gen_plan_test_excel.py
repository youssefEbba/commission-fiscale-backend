# -*- coding: utf-8 -*-
"""
Génère docs/PLAN_DE_TEST.xlsx — fichier de suivi des tests post-livraison SGCI.
Chaque ligne a une colonne "Statut" avec liste déroulante et une colonne "Commentaire".
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import os

OUT = os.path.join(os.path.dirname(__file__), "..", "docs", "PLAN_DE_TEST.xlsx")

# ── Palette couleurs ───────────────────────────────────────────────────────────
C_HEADER_BG   = "1F4E79"   # bleu foncé
C_PHASE_BG    = "2E75B6"   # bleu moyen  (ligne de phase)
C_SECTION_BG  = "D6E4F0"   # bleu très clair (ligne de section)
C_WHITE       = "FFFFFF"
C_GREY_ROW    = "F2F7FB"   # alternance lignes
C_GREEN       = "70AD47"
C_ORANGE      = "ED7D31"
C_RED         = "FF0000"
C_YELLOW      = "FFD966"
C_LIGHT_GREEN = "E2EFDA"
C_LIGHT_RED   = "FFDADA"
C_LIGHT_YELL  = "FFF2CC"

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, bold=True, color=C_WHITE):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def cell_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

# ── Données ────────────────────────────────────────────────────────────────────
# Format : ("PHASE", titre_phase) | ("SECTION", titre_section) | tuple test

ROWS = [
    # ── PHASE 1 ───────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 1 — Infrastructure & Sécurité (Pré-requis)"),
    ("SECTION", "1.1 Démarrage & Santé"),
    ("TEST", "1.1.1", "Health check", "GET /actuator/health", "status: UP"),
    ("TEST", "1.1.2", "Base de données", "Vérifier logs au démarrage", "Aucune erreur Hibernate / migration"),
    ("TEST", "1.1.3", "Migration schéma", "Vérifier colonnes notification.type et entity_type", "VARCHAR(64) minimum"),
    ("TEST", "1.1.4", "Données seed", "Connexion admin → lister utilisateurs", "Comptes DGD, DGI, DGB, DGTCP, Président, AC, Entreprise présents"),
    ("SECTION", "1.2 Authentification"),
    ("TEST", "1.2.1", "Login valide", "POST /api/auth/login avec identifiants corrects", "200 OK + token JWT"),
    ("TEST", "1.2.2", "Login invalide", "Mauvais mot de passe", "401 UNAUTHORIZED"),
    ("TEST", "1.2.3", "Token expiré", "Appel avec JWT expiré", "401 UNAUTHORIZED"),
    ("TEST", "1.2.4", "Sans token", "GET /api/certificats-credit sans Authorization", "401 UNAUTHORIZED"),
    ("TEST", "1.2.5", "/api/auth/me", "Appel avec token valide", "Retourne userId, role, authorities"),
    ("TEST", "1.2.6", "Mauvais rôle", "Entreprise tente filtre admin", "Résultat filtré à son périmètre uniquement"),
    ("SECTION", "1.3 Changement de mot de passe"),
    ("TEST", "1.3.1", "Changement valide", "PATCH /api/utilisateurs/me/password avec ancien MDP correct + nouveau ≥8 chars", "200 OK"),
    ("TEST", "1.3.2", "Ancien MDP incorrect", "Mauvais currentPassword", "400 BAD_REQUEST"),
    ("TEST", "1.3.3", "Nouveau MDP trop court", "newPassword < 8 caractères", "400 BAD_REQUEST"),
    ("TEST", "1.3.4", "Reconnexion", "Se connecter avec le nouveau MDP", "200 OK + token"),

    # ── PHASE 2 ───────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 2 — Référentiels & Administration"),
    ("SECTION", "2.1 Gestion des utilisateurs (ADMIN_SI)"),
    ("TEST", "2.1.1", "Lister utilisateurs", "GET /api/utilisateurs en tant qu'Admin", "Liste complète"),
    ("TEST", "2.1.2", "Créer utilisateur DGD", "POST /api/auth/register avec role: DGD", "201 CREATED"),
    ("TEST", "2.1.3", "Activer compte", "Activer le compte créé", "actif=true"),
    ("TEST", "2.1.4", "Désactiver compte", "Désactiver le compte", "Login refusé"),
    ("TEST", "2.1.5", "Modifier utilisateur", "PATCH informations", "200 OK"),
    ("SECTION", "2.2 Réinitialisation de mot de passe"),
    ("TEST", "2.2.1", "Vérification email", "POST /api/auth/password-reset/check-email avec email connu", "found: true"),
    ("TEST", "2.2.2", "Email inconnu", "Même endpoint avec email inexistant", "found: false"),
    ("TEST", "2.2.3", "Demande de reset", "POST /api/auth/password-reset/request", "202 ACCEPTED + notification DB"),
    ("TEST", "2.2.4", "Traitement Admin", "Admin approuve la demande", "Notification PASSWORD_RESET_TRAITEE créée"),
    ("TEST", "2.2.5", "E-mail reset", "Vérifier boîte mail du demandeur", "E-mail reçu avec les nouvelles informations"),
    ("SECTION", "2.3 Gestion des permissions (ADMIN_SI)"),
    ("TEST", "2.3.1", "Lister permissions", "GET /api/permissions", "Liste complète des permissions"),
    ("TEST", "2.3.2", "Permissions par rôle", "GET /api/permissions/by-role?role=DGD", "Permissions DGD exactes"),
    ("TEST", "2.3.3", "Modifier permission", "Retirer une permission d'un rôle", "Utilisateur perd l'accès (après reconnexion)"),
    ("SECTION", "2.4 Types de documents & Exigences"),
    ("TEST", "2.4.1", "Lister types de docs", "GET /api/referentiel-type-document", "Liste des codes de documents"),
    ("TEST", "2.4.2", "Créer type de doc", "POST avec code + libellé", "201 CREATED"),
    ("TEST", "2.4.3", "Lister exigences", "GET /api/document-requirements", "Exigences par processus"),
    ("TEST", "2.4.4", "Créer exigence", "Ajouter un document obligatoire à un processus", "Document devient requis"),
    ("SECTION", "2.5 Référentiels (Devises, Taux, Bailleurs)"),
    ("TEST", "2.5.1", "Taux forex", "GET /api/forex/rate?from=USD&to=MRU", "Taux retourné"),
    ("TEST", "2.5.2", "Conversion montant", "GET /api/forex/convert?from=USD&to=MRU&amount=1000", "Montant converti"),
    ("TEST", "2.5.3", "Taux interne", "GET /api/taux-change?devise=USD", "Taux MRU retourné"),
    ("TEST", "2.5.4", "Lister devises", "GET /api/devises", "Liste des devises"),
    ("TEST", "2.5.5", "Lister bailleurs", "GET /api/bailleurs", "Liste des bailleurs"),

    # ── PHASE 3 ───────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 3 — Workflow Convention & Marché"),
    ("SECTION", "3.1 Créer une convention (AC)"),
    ("TEST", "3.1.1", "Créer convention", "POST /api/conventions avec bailleur, montant, devise", "201 CREATED + statut EN_ATTENTE"),
    ("TEST", "3.1.2", "Lister conventions AC", "GET /api/conventions en tant qu'AC", "Uniquement ses conventions"),
    ("TEST", "3.1.3", "Consulter détail", "GET /api/conventions/{id}", "Détail complet"),
    ("TEST", "3.1.4", "Téléverser document", "POST /api/conventions/{id}/documents multipart", "Document enregistré"),
    ("TEST", "3.1.5", "Remplacer document", "PUT /api/conventions/{id}/documents/{docId}", "Nouvelle version v2 créée, ancienne inactive"),
    ("TEST", "3.1.6", "Supprimer document", "DELETE /api/conventions/{id}/documents/{docId}", "200 OK"),
    ("SECTION", "3.2 Valider la convention"),
    ("TEST", "3.2.1", "Valider (DGB)", "PATCH /api/conventions/{id}/statut avec statut=VALIDE", "200 OK + statut VALIDE"),
    ("TEST", "3.2.2", "Rejeter (DGB)", "PATCH avec statut=REJETE + motif", "Statut REJETE + motif enregistré"),
    ("TEST", "3.2.3", "Annuler (AC)", "PATCH avec statut=ANNULEE", "Statut ANNULEE"),
    ("TEST", "3.2.4", "Notification", "GET /api/notifications (AC)", "CONVENTION_STATUT_CHANGE reçue"),
    ("TEST", "3.2.5", "E-mail", "Vérifier boîte mail AC", "E-mail de validation/rejet reçu"),
    ("SECTION", "3.3 Gérer les marchés (AC)"),
    ("TEST", "3.3.1", "Créer marché", "POST /api/marches lié à convention validée", "201 CREATED"),
    ("TEST", "3.3.2", "Modifier marché", "PUT /api/marches/{id}", "200 OK"),
    ("TEST", "3.3.3", "Lister marchés", "GET /api/marches", "Liste filtrée à l'AC"),
    ("TEST", "3.3.4", "Vérif correction active", "GET /api/marches/{id}/demande-correction-active", "false si aucune demande"),
    ("TEST", "3.3.5", "Documents marché", "POST /api/marches/{id}/documents multipart", "Document enregistré"),
    ("SECTION", "3.4 Gérer les délégués (AC)"),
    ("TEST", "3.4.1", "Créer délégué", "POST /api/delegues", "201 CREATED"),
    ("TEST", "3.4.2", "Affecter au marché", "PATCH /api/marches/{id}/assign", "Délégué principal affecté"),
    ("TEST", "3.4.3", "Ajouter délégué supp.", "POST /api/marches/{id}/delegues", "201 CREATED"),
    ("TEST", "3.4.4", "Lister marchés délégué", "GET /api/delegues/{id}/marches", "Marchés affectés"),
    ("TEST", "3.4.5", "Désactiver délégué", "PATCH /api/delegues/{id}/actif?actif=false", "Login refusé pour le délégué"),
    ("TEST", "3.4.6", "Retirer délégué", "DELETE /api/marches/{id}/delegues/{delegueId}", "200 OK"),

    # ── PHASE 4 ───────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 4 — Workflow Demande de Correction"),
    ("SECTION", "4.1 Soumission (AC)"),
    ("TEST", "4.1.1", "Créer brouillon", "POST /api/demandes-correction", "201 CREATED + statut BROUILLON"),
    ("TEST", "4.1.2", "Modifier brouillon", "PUT /api/demandes-correction/{id}", "200 OK"),
    ("TEST", "4.1.3", "Joindre offre", "POST /api/demandes-correction/{id}/documents", "Document code OFFRE enregistré"),
    ("TEST", "4.1.4", "Soumettre", "POST /api/demandes-correction/{id}/soumettre", "Statut → RECUE"),
    ("TEST", "4.1.5", "Supprimer brouillon", "DELETE sur brouillon non soumis", "200 OK"),
    ("TEST", "4.1.6", "Supprimer soumis", "Tenter DELETE sur demande RECUE", "403 FORBIDDEN"),
    ("SECTION", "4.2 Circuit DGD"),
    ("TEST", "4.2.1", "Visa DGD + offre corrigée", "DGD : Visa sans OFFRE_FISCALE_CORRIGEE → modale upload → visa", "Document actif + décision VISA DGD"),
    ("TEST", "4.2.2", "Rejet définitif DGD", "decision=REJET_DEFINITIF", "Statut → REJETEE"),
    ("TEST", "4.2.3", "Rejet temporaire DGD", "decision=REJET_TEMPORAIRE avec motif", "Statut → INCOMPLETE"),
    ("TEST", "4.2.4", "Réponse AC", "POST .../decisions/{decId}/rejet-temp/reponses", "Réponse enregistrée"),
    ("TEST", "4.2.5", "Résolution", "PUT .../decisions/{decId}/resolve", "Statut → A_RECONTROLER"),
    ("TEST", "4.2.6", "Notification", "GET /api/notifications (AC)", "REJET_TEMP_DECISION + REJET_TEMP_RESOLU reçus"),
    ("SECTION", "4.3 Circuit DGTCP → DGI → DGB"),
    ("TEST", "4.3.1", "Visa DGTCP", "decision=VISA (DGTCP)", "Décision enregistrée"),
    ("TEST", "4.3.2", "Visa DGI + upload crédit intérieur", "DGI : Visa sans CREDIT_INTERIEUR → modale upload → visa", "Document CREDIT_INTERIEUR actif + décision VISA DGI"),
    ("TEST", "4.3.3", "Visa DGB", "decision=VISA (DGB)", "Décision enregistrée"),
    ("TEST", "4.3.4", "Historique décisions", "GET /api/demandes-correction/{id}/decisions", "3 visas listés"),
    ("TEST", "4.3.5", "Garde upload DGI", "DGI tente POST .../documents?codeDocument=LETTRE_SAISINE", "403 — seul CREDIT_INTERIEUR autorisé"),
    ("TEST", "4.3.6", "Garde visa DGI", "DGI tente decision=VISA sans document actif", "400 — document CREDIT_INTERIEUR requis"),
    ("SECTION", "4.4 Décision Président"),
    ("TEST", "4.4.1", "Adopter", "PATCH /api/demandes-correction/{id}/statut?statut=ADOPTEE", "Statut ADOPTEE"),
    ("TEST", "4.4.2", "Générer lettre", "Action de génération de lettre", "Document lettre généré"),
    ("TEST", "4.4.3", "Déposer signature", "Upload document signé", "Document signé enregistré"),
    ("TEST", "4.4.4", "Notifier", "statut=NOTIFIEE", "Statut NOTIFIEE"),
    ("TEST", "4.4.5", "Rejeter", "statut=REJETEE", "Statut REJETEE"),
    ("TEST", "4.4.6", "Notifications", "Vérifier AC + Entreprise", "CORRECTION_ADOPTEE reçu"),
    ("TEST", "4.4.7", "E-mail adoption", "Vérifier boîte mail", "E-mails envoyés aux destinataires"),
    ("SECTION", "4.5 Réclamations (après adoption)"),
    ("TEST", "4.5.1", "Déposer réclamation", "POST /api/demandes-correction/{id}/reclamations (Entreprise)", "201 CREATED"),
    ("TEST", "4.5.2", "Lister réclamations", "GET /api/demandes-correction/{id}/reclamations", "Réclamation listée"),
    ("TEST", "4.5.3", "Annuler réclamation", "POST .../reclamations/{recId}/annuler (avant traitement)", "200 OK"),
    ("TEST", "4.5.4", "Accepter réclamation", "PATCH .../reclamations/{recId} avec acceptee=true (DGTCP)", "200 OK"),
    ("TEST", "4.5.5", "Rejeter réclamation", "acceptee=false + motif", "200 OK"),
    ("SECTION", "4.6 Demandes d'explication"),
    ("TEST", "4.6.1", "Créer explication", "POST /api/demandes-explication (DGD) avec contexte=CORRECTION", "201 CREATED"),
    ("TEST", "4.6.2", "Répondre", "POST /api/demandes-explication/{id}/messages (AC)", "Message ajouté"),
    ("TEST", "4.6.3", "Fermer", "PUT /api/demandes-explication/{id}/fermer", "Statut FERMEE"),

    # ── PHASE 5 ───────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 5 — Workflow Mise en place du Certificat"),
    ("SECTION", "5.1 Création & soumission (AC)"),
    ("TEST", "5.1.1", "Créer certificat", "POST /api/certificats-credit lié à demande ADOPTEE", "201 CREATED + statut ENVOYEE"),
    ("TEST", "5.1.2", "Modifier brouillon", "PUT /api/certificats-credit/{id}", "200 OK"),
    ("TEST", "5.1.3", "Joindre documents", "POST /api/certificats-credit/{id}/documents", "Documents requis joints"),
    ("TEST", "5.1.4", "Supprimer brouillon", "DELETE /api/certificats-credit/{id}", "200 OK"),
    ("TEST", "5.1.5", "Double soumission", "Créer 2 certificats sur la même demande", "409 ou erreur métier"),
    ("SECTION", "5.2 Contrôle & Visa (DGD / DGI / DGTCP)"),
    ("TEST", "5.2.1", "Prise en charge DGD", "POST /api/certificats-credit/{id}/prendre-en-charge", "Statut → EN_CONTROLE"),
    ("TEST", "5.2.2", "Définir montants DGTCP", "PATCH /api/certificats-credit/{id}/montants", "Montants enregistrés"),
    ("TEST", "5.2.3", "Visa DGI", "POST .../decisions avec decision=VISA", "201 CREATED"),
    ("TEST", "5.2.4", "Visa DGD", "POST .../decisions avec decision=VISA", "201 CREATED"),
    ("TEST", "5.2.5", "Visa DGTCP", "POST .../decisions avec decision=VISA", "201 CREATED + statut EN_VALIDATION_PRESIDENT"),
    ("TEST", "5.2.6", "Rejet temporaire DGD", "decision=REJET_TEMPORAIRE", "Statut → INCOMPLETE"),
    ("TEST", "5.2.7", "Réponse entreprise", "Upload compléments", "Documents ajoutés"),
    ("TEST", "5.2.8", "Résolution rejet", "PUT .../decisions/{decId}/resolve", "Statut → A_RECONTROLER"),
    ("SECTION", "5.3 Validation Président & Ouverture DGTCP"),
    ("TEST", "5.3.1", "Valider (Président)", "PATCH .../statut?statut=VALIDE_PRESIDENT", "Statut → VALIDE_PRESIDENT"),
    ("TEST", "5.3.2", "Générer document", "Génération du certificat signé", "Document PDF généré"),
    ("TEST", "5.3.3", "Ouvrir crédit DGTCP", "PATCH .../statut?statut=OUVERT (DGTCP)", "Statut → OUVERT + soldeCordon initialisé"),
    ("TEST", "5.3.4", "Vérifier soldes", "GET /api/certificats-credit/{id}", "soldeCordon = montantCordon, soldeTVA = montantTVAInterieure"),
    ("TEST", "5.3.5", "Stock TVA", "GET /api/certificats-credit/{id}/tva-stock", "Stock TVA déductible initialisé"),
    ("TEST", "5.3.6", "Notification", "Vérifier notifications Entreprise", "CERTIFICAT_OUVERT reçu"),
    ("TEST", "5.3.7", "Annuler", "statut=ANNULE (Président)", "Statut → ANNULE, soldes à 0"),

    # ── PHASE 6 ───────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 6 — Workflow Utilisation Douane"),
    ("SECTION", "6.1 Création & soumission (Entreprise)"),
    ("TEST", "6.1.1", "Créer utilisation", "POST /api/utilisations-credit avec type=DOUANIER", "201 CREATED + statut DEMANDEE"),
    ("TEST", "6.1.2", "Certificat fermé", "Tenter utilisation sur certificat CLOTURE", "Erreur métier"),
    ("TEST", "6.1.3", "Montant > solde", "montantDroits > soldeCordon du certificat", "Erreur métier / 400"),
    ("TEST", "6.1.4", "Solde = 0", "Tenter utilisation avec solde à 0", "Erreur métier"),
    ("TEST", "6.1.5", "Modifier brouillon", "PUT /api/utilisations-credit/{id}", "200 OK"),
    ("TEST", "6.1.6", "Supprimer brouillon", "DELETE /api/utilisations-credit/{id}", "200 OK"),
    ("TEST", "6.1.7", "Documents", "POST .../documents (BL, déclaration SYDONIA)", "Documents enregistrés"),
    ("SECTION", "6.2 Contrôle DGD & Visa bulletin"),
    ("TEST", "6.2.1", "Prise en charge", "PATCH .../statut?statut=EN_VERIFICATION", "Statut → EN_VERIFICATION"),
    ("TEST", "6.2.2", "Rejet définitif", "PATCH .../statut?statut=REJETEE", "Statut → REJETEE, solde non débité"),
    ("TEST", "6.2.3", "Rejet temporaire", "POST .../decisions avec REJET_TEMPORAIRE", "Statut → INCOMPLETE"),
    ("TEST", "6.2.4", "Réponse entreprise", "Upload compléments", "Documents ajoutés"),
    ("TEST", "6.2.5", "Résolution", "PUT .../decisions/{decId}/resolve", "Statut → A_RECONTROLER"),
    ("TEST", "6.2.6", "Visa DGD bulletin", "POST /api/utilisations-credit/{id}/visa-dgd avec annotations", "Lignes AU_CI / A_PAYER mises à jour"),
    ("TEST", "6.2.7", "Consulter lignes", "GET /api/utilisations-credit/{id}/lignes-bulletin", "Lignes avec montants DGD"),
    ("TEST", "6.2.8", "Notification visa", "Vérifier notifications Entreprise", "UTIL_DOUANE_VISA_DGD reçu"),
    ("SECTION", "6.3 Saisie chèque & Envoi Trésor"),
    ("TEST", "6.3.1", "Saisir chèque", "POST /api/utilisations-credit/{id}/cheque (Entreprise)", "Statut → CHEQUE_SAISI"),
    ("TEST", "6.3.2", "Chèque sans visa", "Tenter saisie avant visa DGD", "Erreur métier"),
    ("TEST", "6.3.3", "Envoyer au Trésor", "POST /api/utilisations-credit/{id}/envoyer-au-tresor (DGTCP)", "Statut → ENVOYEE_AU_TRESOR"),
    ("TEST", "6.3.4", "Saisir quittances", "POST /api/utilisations-credit/{id}/quittances (DGTCP)", "Statut → QUITTANCES_ENREGISTREES"),
    ("TEST", "6.3.5", "Consulter quittances", "GET /api/utilisations-credit/{id}/quittances", "Quittances listées"),
    ("SECTION", "6.4 Liquidation & Clôture"),
    ("TEST", "6.4.1", "Liquidation douane", "POST /api/utilisations-credit/{id}/liquidation-douane (DGTCP)", "Statut → LIQUIDEE + solde cordon débité"),
    ("TEST", "6.4.2", "Vérifier solde", "GET /api/certificats-credit/{id}", "soldeCordon = ancienSolde − montantLiquidé"),
    ("TEST", "6.4.3", "Solde négatif", "Vérifier solde ≥ 0 après liquidation", "Impossible si montant > solde"),
    ("TEST", "6.4.4", "Accuser réception", "POST /api/utilisations-credit/{id}/cloturer-reception (Entreprise)", "Statut → CLOTUREE"),
    ("TEST", "6.4.5", "Notification clôture", "Vérifier notifications Entreprise", "UTIL_DOUANE_CLOTUREE reçu"),
    ("TEST", "6.4.6", "E-mail liquidation", "Vérifier boîte mail", "E-mail de liquidation reçu"),

    # ── PHASE 7 ───────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 7 — Workflow Utilisation TVA Intérieure"),
    ("TEST", "7.1",  "Créer utilisation TVA", "POST /api/utilisations-credit avec type=TVA_INTERIEURE", "201 CREATED + statut DEMANDEE"),
    ("TEST", "7.2",  "Montant > soldeTVA", "Montant > soldeTVA du certificat", "Erreur métier"),
    ("TEST", "7.3",  "TVA déductible stock", "GET /api/certificats-credit/{id}/tva-stock", "Stock TVA disponible retourné"),
    ("TEST", "7.4",  "Instruction DGI", "POST .../decisions avec decision=VISA (DGI)", "Décision enregistrée"),
    ("TEST", "7.5",  "Rejet temporaire DGI", "REJET_TEMPORAIRE + motif", "Statut → INCOMPLETE"),
    ("TEST", "7.6",  "Réponse + résolution", "Upload + resolve", "Statut → A_RECONTROLER"),
    ("TEST", "7.7",  "Vérification DGTCP", "PATCH .../statut?statut=EN_VERIFICATION", "Statut → EN_VERIFICATION"),
    ("TEST", "7.8",  "Validation DGTCP", "statut=VISE", "Statut → VISE"),
    ("TEST", "7.9",  "Apurement TVA", "POST /api/utilisations-credit/{id}/apurement-tva", "Statut → APUREE + soldeTVA débité"),
    ("TEST", "7.10", "Vérifier soldeTVA", "GET /api/certificats-credit/{id}", "soldeTVA = ancien − montant apuré"),
    ("TEST", "7.11", "Rejet définitif", "statut=REJETEE", "Solde non débité"),
    ("TEST", "7.12", "Notification", "GET /api/notifications (Entreprise)", "UTIL_TVA_APUREE reçu"),

    # ── PHASE 8 ───────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 8 — Workflow Transfert de Crédit"),
    ("TEST", "8.1",  "Créer transfert", "POST /api/transferts-credit", "201 CREATED + statut DEMANDE"),
    ("TEST", "8.2",  "Déposer pièces", "POST /api/transferts-credit/{id}/documents", "Statut → EN_COURS automatique"),
    ("TEST", "8.3",  "Lister transferts (DGTCP)", "GET /api/transferts-credit", "File DGTCP visible"),
    ("TEST", "8.4",  "Validation DGTCP", "POST /api/transferts-credit/{id}/valider", "Statut → TRANSFERE + soldes réaffectés"),
    ("TEST", "8.5",  "Vérifier soldes", "GET /api/certificats-credit/{id}", "Soldes mis à jour après transfert"),
    ("TEST", "8.6",  "Rejet définitif", "POST /api/transferts-credit/{id}/rejeter", "Statut → REJETE"),
    ("TEST", "8.7",  "Rejet temporaire", "POST .../decisions REJET_TEMPORAIRE", "Statut → INCOMPLETE"),
    ("TEST", "8.8",  "Réponse entreprise", "Upload compléments", "Réponse enregistrée"),
    ("TEST", "8.9",  "Résolution", "PUT .../decisions/{decId}/resolve", "Statut → A_RECONTROLER"),
    ("TEST", "8.10", "Validation Président", "POST /api/transferts-credit/{id}/valider", "200 OK"),
    ("TEST", "8.11", "Annulation", "POST /api/transferts-credit/{id}/annuler", "Statut → ANNULEE"),
    ("TEST", "8.12", "Annuler après exec.", "Annuler un transfert TRANSFERE", "Erreur métier"),
    ("TEST", "8.13", "Notification", "GET /api/notifications (Entreprise)", "TRANSFERT_VALIDE / TRANSFERT_REJETE reçu"),

    # ── PHASE 9 ───────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 9 — Workflow Sous-traitance"),
    ("TEST", "9.1",  "Onboarding", "POST /api/sous-traitances/onboarding", "Compte sous-traitant + dossier créés"),
    ("TEST", "9.2",  "Soumettre demande", "POST /api/sous-traitances", "201 CREATED + statut DEMANDE"),
    ("TEST", "9.3",  "Lister (DGTCP)", "GET /api/sous-traitances", "File DGTCP visible"),
    ("TEST", "9.4",  "Autoriser", "POST /api/sous-traitances/{id}/autoriser", "Statut → AUTORISEE"),
    ("TEST", "9.5",  "Refuser", "POST /api/sous-traitances/{id}/refuser", "Statut → REFUSEE"),
    ("TEST", "9.6",  "Suspendre", "POST .../suspendre-titulaire", "Statut → SUSPENDUE"),
    ("TEST", "9.7",  "Réactiver", "POST .../reactiver-titulaire", "Statut → AUTORISEE"),
    ("TEST", "9.8",  "Révoquer", "POST .../revoquer-titulaire", "Statut → REVOQUEE (irréversible)"),
    ("TEST", "9.9",  "Accès sous-traitant", "GET /api/certificats-credit", "Voit uniquement le certificat partagé"),
    ("TEST", "9.10", "Utilisation ST", "Créer utilisation sur certificat partagé", "Utilisation dans les limites du solde alloué"),
    ("TEST", "9.11", "Lister entreprises ST", "GET /api/sous-traitances/entreprises-sous-traitantes", "Liste des sous-traitants du titulaire"),

    # ── PHASE 10 ──────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 10 — Workflow Clôture du Certificat"),
    ("TEST", "10.1",  "File éligibles", "GET /api/clotures-credit/queue", "Certificats avec motifs de blocage"),
    ("TEST", "10.2",  "Certificats clôturables", "GET /api/clotures-credit/eligible", "IDs des certificats immédiatement clôturables"),
    ("TEST", "10.3",  "Proposer clôture", "POST /api/clotures-credit", "Proposition créée"),
    ("TEST", "10.4",  "Documents clôture", "POST /api/clotures-credit/{id}/documents", "Documents joints"),
    ("TEST", "10.5",  "File Président", "GET /api/clotures-credit/propositions", "Proposition visible"),
    ("TEST", "10.6",  "Valider clôture", "POST /api/clotures-credit/{id}/valider", "200 OK"),
    ("TEST", "10.7",  "Rejeter clôture", "POST /api/clotures-credit/{id}/rejeter", "200 OK"),
    ("TEST", "10.8",  "Finaliser", "POST /api/clotures-credit/{id}/finaliser", "Certificat → CLOTURE"),
    ("TEST", "10.9",  "Vérif post-clôture", "Tenter utilisation sur certificat CLOTURE", "Erreur métier"),
    ("TEST", "10.10", "Notification", "GET /api/notifications (Entreprise)", "CLOTURE_FINALISEE reçu + e-mail"),

    # ── PHASE 11 ──────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 11 — GED & Documents (versionnage)"),
    ("TEST", "11.1", "Upload v1", "POST .../documents (document quelconque)", "Version 1 créée, actif=true"),
    ("TEST", "11.2", "Remplacer → v2", "PUT .../documents/{id} avec nouveau fichier", "Version 2 créée, v1 passe à actif=false"),
    ("TEST", "11.3", "Versions conservées", "Consulter dossier GED", "V1 et V2 présentes (v1 inactive)"),
    ("TEST", "11.4", "Dossier GED", "GET /api/dossiers", "Dossiers accessibles selon rôle"),
    ("TEST", "11.5", "Dossier détail", "GET /api/dossiers/{id}", "Arborescence complète avec versions"),
    ("TEST", "11.6", "Admin config types", "Créer type doc + exigence", "Document devient obligatoire dans le processus ciblé"),

    # ── PHASE 12 ──────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 12 — Notifications & Commission Relais"),
    ("SECTION", "12.1 Notifications in-app"),
    ("TEST", "12.1.1", "Lister notifications", "GET /api/notifications", "Notifications de la session"),
    ("TEST", "12.1.2", "Compteur non lus", "GET /api/notifications/unread-count", "Nombre correct"),
    ("TEST", "12.1.3", "Marquer lue", "PATCH /api/notifications/{id}/read", "isRead=true"),
    ("TEST", "12.1.4", "Tout marquer lu", "PATCH /api/notifications/read-all", "Toutes isRead=true"),
    ("SECTION", "12.2 E-mails"),
    ("TEST", "12.2.1", "Correction soumise", "Déclencher soumission → vérifier boîte mail", "Président + DGD + membres Commission"),
    ("TEST", "12.2.2", "Correction adoptée", "Déclencher adoption → vérifier boîte mail", "AC + Entreprise"),
    ("TEST", "12.2.3", "Certificat ouvert", "Déclencher ouverture → vérifier boîte mail", "Entreprise"),
    ("TEST", "12.2.4", "Utilisation liquidée", "Déclencher liquidation → vérifier boîte mail", "Entreprise"),
    ("TEST", "12.2.5", "Transfert validé/rejeté", "Déclencher décision → vérifier boîte mail", "Entreprise"),
    ("TEST", "12.2.6", "Clôture finalisée", "Déclencher clôture → vérifier boîte mail", "Entreprise"),
    ("TEST", "12.2.7", "Reset MDP approuvé", "Approuver reset → vérifier boîte mail", "Utilisateur demandeur"),
    ("SECTION", "12.3 Commission Relais (impersonation)"),
    ("TEST", "12.3.1", "Lister entreprises", "GET /api/commission-relais/entreprises", "Liste paginée"),
    ("TEST", "12.3.2", "Impersonate entreprise", "POST /api/commission-relais/impersonate/entreprise", "Nouveau JWT rôle entreprise"),
    ("TEST", "12.3.3", "Action impersonée", "Créer utilisation avec le JWT impersoné", "201 CREATED"),
    ("TEST", "12.3.4", "Release", "POST /api/commission-relais/release", "JWT Commission Relais restauré"),
    ("TEST", "12.3.5", "Impersonate AC", "POST /api/commission-relais/impersonate/autorite-contractante", "Nouveau JWT rôle AC"),

    # ── PHASE 13 ──────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 13 — Reporting & Audit"),
    ("TEST", "13.1", "Synthèse nationale", "GET /api/reporting/summary (Président)", "Agrégats complets toutes AC"),
    ("TEST", "13.2", "Synthèse filtrée", "GET .../summary?entrepriseId=X (DGTCP)", "Données filtrées à l'entreprise X"),
    ("TEST", "13.3", "Synthèse AC", "GET /api/reporting/summary (AC)", "Données limitées à son périmètre"),
    ("TEST", "13.4", "Synthèse Entreprise", "GET /api/reporting/summary (Entreprise)", "Uniquement ses propres dossiers"),
    ("TEST", "13.5", "Série temporelle", "GET /api/reporting/timeseries/demandes (Président)", "Points mensuels"),
    ("TEST", "13.6", "Période filtrée", "?from=2026-01-01T00:00:00Z&to=2026-06-01T00:00:00Z", "Données dans la fenêtre"),
    ("TEST", "13.7", "Journal audit", "GET /api/audit-logs (Admin)", "Actions sensibles tracées"),
    ("TEST", "13.8", "Snapshot objet", "Consulter un log", "objectSnapshot JSON présent"),

    # ── PHASE 14 ──────────────────────────────────────────────────────────────
    ("PHASE", "PHASE 14 — Tests de régression & Cas limites"),
    ("TEST", "14.1",  "Solde exact", "Montant utilisation = solde exact restant", "Accepté, solde → 0"),
    ("TEST", "14.2",  "Double soumission", "Soumettre deux fois le même brouillon", "Erreur à la 2e tentative"),
    ("TEST", "14.3",  "Workflow hors ordre", "DGTCP tente de viser avant DGI", "Erreur métier"),
    ("TEST", "14.4",  "Certificat non ouvert", "Utilisation sur certificat EN_CONTROLE", "Erreur métier"),
    ("TEST", "14.5",  "Upload fichier vide", "POST .../documents avec fichier 0 bytes", "400 BAD_REQUEST"),
    ("TEST", "14.6",  "Token autre user", "Utiliser token de l'entreprise A sur ressources de B", "403 FORBIDDEN"),
    ("TEST", "14.7",  "Pagination reporting", "?from= très ancienne date", "Résultats corrects sans erreur"),
    ("TEST", "14.8",  "Mail timeout", "Serveur SMTP indisponible", "Transaction métier réussie, mail en échec logué uniquement"),
    ("TEST", "14.9",  "Rejet temp x2", "Deux rejets temporaires ouverts simultanément", "Les deux doivent être résolus avant de continuer"),
    ("TEST", "14.10", "Solde négatif impossible", "Liquidation > solde restant", "Rejet avec message d'erreur clair"),
    ("TEST", "14.11", "Révocation sous-traitance", "Tenter utilisation après révocation", "Accès refusé"),
    ("TEST", "14.12", "Demande explication fermée", "Envoyer un message sur demande FERMEE", "400 BAD_REQUEST"),
]

# ── Descriptions "Comment tester (UI)" ─────────────────────────────────────────
# Parcours fonctionnel côté front (cs-front-for-cursor). Menus : Tableau de bord,
# Conventions, Marchés, Demandes (> Correction / Mise en place), Représentants,
# Certificats, Utilisations, Simulation, Reporting, Opérations (> Modifications /
# Transferts / Sous-traitance / Clôture), GED, Paramétrage (Utilisateurs, Rôles, Audit).
DESCRIPTIONS = {
    # Phase 1
    "1.1.1": "Ouvrir l'URL de l'application dans le navigateur : la page de connexion doit s'afficher.",
    "1.1.2": "Au lancement du backend (terminal), parcourir les logs de démarrage : aucune erreur ne doit apparaître.",
    "1.1.3": "Via un outil SQL, vérifier la structure des colonnes de la table notification après démarrage.",
    "1.1.4": "Se connecter en admin → menu Paramétrage > Utilisateurs : la liste des comptes s'affiche.",
    "1.2.1": "Page de connexion : saisir identifiant + mot de passe valides, cliquer sur Se connecter → redirection vers le tableau de bord.",
    "1.2.2": "Page de connexion : saisir un mauvais mot de passe → message d'erreur, aucune connexion.",
    "1.2.3": "Attendre l'expiration du jeton (ou l'altérer) : au clic suivant l'app redirige vers la page de connexion.",
    "1.2.4": "Sans être connecté, coller une URL du tableau de bord → redirection automatique vers /login.",
    "1.2.5": "Une fois connecté, le nom et le rôle s'affichent en haut à droite (menu profil).",
    "1.2.6": "Se connecter en Entreprise : seuls ses propres dossiers apparaissent dans les listes.",
    "1.3.1": "Menu profil (haut à droite) > changer le mot de passe : saisir l'ancien + le nouveau (≥8), valider → succès.",
    "1.3.2": "Même écran : saisir un ancien mot de passe incorrect → message d'erreur.",
    "1.3.3": "Même écran : saisir un nouveau mot de passe trop court → message de validation.",
    "1.3.4": "Se déconnecter puis se reconnecter avec le nouveau mot de passe.",
    # Phase 2
    "2.1.1": "Admin → Paramétrage > Utilisateurs : la liste complète s'affiche.",
    "2.1.2": "Écran d'inscription/création de compte : créer un utilisateur avec le rôle DGD.",
    "2.1.3": "Paramétrage > Utilisateurs : activer un compte en attente (bascule Actif).",
    "2.1.4": "Paramétrage > Utilisateurs : désactiver un compte puis vérifier que la connexion est refusée.",
    "2.1.5": "Paramétrage > Utilisateurs : ouvrir un utilisateur, modifier nom/e-mail, enregistrer.",
    "2.2.1": "Page de connexion > Mot de passe oublié : saisir un e-mail connu → message de transmission.",
    "2.2.2": "Même écran : saisir un e-mail inconnu → message générique (pas de fuite d'information).",
    "2.2.3": "Soumettre la demande de réinitialisation → confirmation affichée.",
    "2.2.4": "Admin → écran des demandes de réinitialisation : approuver une demande en attente.",
    "2.2.5": "Consulter la boîte mail de l'utilisateur : e-mail reçu avec le nouveau mot de passe.",
    "2.3.1": "Admin → Paramétrage > Rôles & permissions : la liste des permissions s'affiche.",
    "2.3.2": "Sélectionner le rôle DGD : ses permissions s'affichent.",
    "2.3.3": "Retirer une permission d'un rôle, enregistrer ; l'utilisateur perd l'accès après reconnexion.",
    "2.4.1": "Admin → GED > Configuration : la liste des types de documents s'affiche.",
    "2.4.2": "GED > Configuration : ajouter un type de document (code + libellé).",
    "2.4.3": "GED > Configuration : consulter les exigences documentaires par processus.",
    "2.4.4": "GED > Configuration : marquer un document comme obligatoire pour un processus.",
    "2.5.1": "Écran Simulation / Référentiel taxes : vérifier l'affichage d'un taux de change.",
    "2.5.2": "Simulation : convertir un montant entre deux devises.",
    "2.5.3": "Paramétrage > Référentiel taxes : consulter les taux internes.",
    "2.5.4": "Référentiel taxes : la liste des devises s'affiche.",
    "2.5.5": "Conventions > Nouvelle : la liste des bailleurs est proposée dans le formulaire.",
    # Phase 3
    "3.1.1": "AC → Conventions > Nouvelle convention : renseigner bailleur, montant, devise, enregistrer.",
    "3.1.2": "Menu Conventions : seules les conventions de l'AC connectée sont listées.",
    "3.1.3": "Conventions : cliquer sur une ligne pour ouvrir le détail.",
    "3.1.4": "Détail convention > Documents : téléverser un fichier.",
    "3.1.5": "Détail convention : remplacer un document existant → nouvelle version créée.",
    "3.1.6": "Détail convention : supprimer un document joint.",
    "3.2.1": "DGB → Conventions : ouvrir une convention et cliquer sur Valider.",
    "3.2.2": "DGB → Conventions : rejeter une convention en saisissant un motif.",
    "3.2.3": "AC → Conventions : annuler une convention.",
    "3.2.4": "Vérifier la cloche de notifications (AC) après la décision.",
    "3.2.5": "Consulter la boîte mail de l'AC.",
    "3.3.1": "AC → Marchés > Nouveau marché : rattacher à une convention validée.",
    "3.3.2": "Marchés : ouvrir un marché et modifier ses informations.",
    "3.3.3": "Marchés : la liste filtrée à l'AC s'affiche.",
    "3.3.4": "Détail marché : vérifier l'indicateur de demande de correction active.",
    "3.3.5": "Détail marché : téléverser un document.",
    "3.4.1": "AC → Représentants > Nouveau : créer un délégué.",
    "3.4.2": "Détail marché : affecter le délégué principal.",
    "3.4.3": "Détail marché : ajouter un délégué supplémentaire.",
    "3.4.4": "Représentants : consulter les marchés affectés à un délégué.",
    "3.4.5": "Représentants : désactiver un délégué puis vérifier que sa connexion est refusée.",
    "3.4.6": "Détail marché : retirer un délégué.",
    # Phase 4
    "4.1.1": "AC → Demandes > Demandes de correction > Nouvelle (assistant) : créer un brouillon.",
    "4.1.2": "Demandes de correction : ouvrir le brouillon et le modifier.",
    "4.1.3": "Détail demande : joindre l'offre (document).",
    "4.1.4": "Détail demande : cliquer sur Soumettre.",
    "4.1.5": "Demandes : supprimer un brouillon non soumis.",
    "4.1.6": "Tenter de supprimer une demande déjà soumise → action refusée.",
    "4.2.1": "DGD → ouvrir la demande : cliquer Visa ; si OFFRE_FISCALE_CORRIGEE absent, modale upload → valider → visa enregistré.",
    "4.2.2": "DGD : rejeter définitivement la demande.",
    "4.2.3": "DGD : rejet temporaire avec motif → statut Incomplète.",
    "4.2.4": "AC → détail demande : répondre au rejet temporaire (compléments).",
    "4.2.5": "DGD : marquer le rejet temporaire comme résolu.",
    "4.2.6": "AC : vérifier la cloche de notifications.",
    "4.3.1": "DGTCP → Demandes de correction : apposer le Visa.",
    "4.3.2": "DGI → ouvrir une demande (visa DGD posé) : cliquer Visa ; si CREDIT_INTERIEUR absent, modale upload → valider → visa enregistré.",
    "4.3.3": "DGB → apposer le Visa.",
    "4.3.4": "Détail demande : l'historique des 3 visas s'affiche.",
    "4.3.5": "API ou outil : en DGI, tenter upload LETTRE_SAISINE → 403.",
    "4.3.6": "API : en DGI, tenter visa sans CREDIT_INTERIEUR actif → 400.",
    "4.4.1": "Président → Demandes de correction : Adopter la demande.",
    "4.4.2": "Président : générer la lettre de décision.",
    "4.4.3": "Président : déposer le document signé.",
    "4.4.4": "Président : passer la demande au statut Notifiée.",
    "4.4.5": "Président : rejeter la demande.",
    "4.4.6": "AC et Entreprise : vérifier les notifications.",
    "4.4.7": "Vérifier les boîtes mail des destinataires.",
    "4.5.1": "Entreprise → détail demande adoptée : déposer une réclamation.",
    "4.5.2": "Détail demande : la réclamation s'affiche dans la liste.",
    "4.5.3": "Entreprise : annuler une réclamation avant traitement.",
    "4.5.4": "DGTCP : accepter une réclamation.",
    "4.5.5": "DGTCP : rejeter une réclamation avec motif.",
    "4.6.1": "DGD → détail demande : créer une demande d'explication.",
    "4.6.2": "AC → répondre à la demande d'explication (message).",
    "4.6.3": "Fermer la demande d'explication.",
    # Phase 5
    "5.1.1": "AC → Demandes > Demandes de mise en place > Nouveau : créer le certificat lié à une demande adoptée.",
    "5.1.2": "Mise en place : modifier le brouillon.",
    "5.1.3": "Détail mise en place : joindre les documents requis.",
    "5.1.4": "Mise en place : supprimer le brouillon.",
    "5.1.5": "Tenter de créer un 2e certificat sur la même demande → refusé.",
    "5.2.1": "DGD → détail mise en place : Prendre en charge.",
    "5.2.2": "DGTCP : saisir les montants (cordon + TVA intérieure).",
    "5.2.3": "DGI : apposer le Visa.",
    "5.2.4": "DGD : apposer le Visa.",
    "5.2.5": "DGTCP : apposer le Visa → passage en validation président.",
    "5.2.6": "DGD : rejet temporaire.",
    "5.2.7": "Entreprise : déposer des compléments.",
    "5.2.8": "DGD : marquer le rejet résolu.",
    "5.3.1": "Président → détail mise en place : Valider.",
    "5.3.2": "Générer le certificat signé (PDF).",
    "5.3.3": "DGTCP : Ouvrir le crédit.",
    "5.3.4": "Détail certificat : vérifier les soldes initialisés.",
    "5.3.5": "Détail certificat : vérifier le stock TVA.",
    "5.3.6": "Entreprise : vérifier la notification.",
    "5.3.7": "Président : annuler un certificat.",
    # Phase 6
    "6.1.1": "Entreprise → Utilisations > Nouvelle (type Douanier) : créer la demande.",
    "6.1.2": "Tenter une utilisation sur un certificat clôturé → refusé.",
    "6.1.3": "Saisir un montant supérieur au solde cordon → refusé.",
    "6.1.4": "Tenter une utilisation alors que le solde est à 0 → refusé.",
    "6.1.5": "Utilisations : modifier un brouillon.",
    "6.1.6": "Utilisations : supprimer un brouillon.",
    "6.1.7": "Détail utilisation : joindre le BL et la déclaration SYDONIA.",
    "6.2.1": "DGD → détail utilisation : passer en vérification.",
    "6.2.2": "DGD : rejeter définitivement.",
    "6.2.3": "DGD : rejet temporaire.",
    "6.2.4": "Entreprise : déposer des compléments.",
    "6.2.5": "DGD : marquer le rejet résolu.",
    "6.2.6": "DGD : viser le bulletin (annoter les lignes AU_CI / A_PAYER).",
    "6.2.7": "Détail utilisation : consulter les lignes du bulletin.",
    "6.2.8": "Entreprise : vérifier la notification de visa.",
    "6.3.1": "Entreprise → détail utilisation : saisir le chèque.",
    "6.3.2": "Tenter de saisir le chèque avant le visa DGD → refusé.",
    "6.3.3": "DGTCP : envoyer au Trésor.",
    "6.3.4": "DGTCP : saisir les quittances.",
    "6.3.5": "Détail utilisation : consulter les quittances.",
    "6.4.1": "DGTCP : effectuer la liquidation douane → solde débité.",
    "6.4.2": "Détail certificat : vérifier le solde après liquidation.",
    "6.4.3": "Vérifier que le solde ne devient jamais négatif.",
    "6.4.4": "Entreprise : accuser réception (clôture de l'utilisation).",
    "6.4.5": "Entreprise : vérifier la notification de clôture.",
    "6.4.6": "Consulter la boîte mail.",
    # Phase 7
    "7.1": "Entreprise → Utilisations > Nouvelle (type TVA intérieure).",
    "7.2": "Saisir un montant supérieur au solde TVA → refusé.",
    "7.3": "Détail certificat : consulter le stock TVA disponible.",
    "7.4": "DGI → détail utilisation : instruire / apposer le visa.",
    "7.5": "DGI : rejet temporaire avec motif.",
    "7.6": "Entreprise : compléter, puis DGI : marquer résolu.",
    "7.7": "DGTCP : passer en vérification.",
    "7.8": "DGTCP : valider (statut Visé).",
    "7.9": "DGTCP : apurement TVA → solde débité.",
    "7.10": "Détail certificat : vérifier le solde TVA après apurement.",
    "7.11": "DGTCP : rejeter définitivement.",
    "7.12": "Entreprise : vérifier la notification.",
    # Phase 8
    "8.1": "Entreprise → Opérations > Transferts > Nouveau.",
    "8.2": "Détail transfert : déposer les pièces → passage En cours.",
    "8.3": "DGTCP → Transferts : la file s'affiche.",
    "8.4": "DGTCP : valider le transfert → soldes réaffectés.",
    "8.5": "Détail certificat : vérifier les soldes.",
    "8.6": "DGTCP : rejeter définitivement.",
    "8.7": "DGTCP : rejet temporaire.",
    "8.8": "Entreprise : déposer des compléments.",
    "8.9": "DGTCP : marquer le rejet résolu.",
    "8.10": "Président : valider le transfert.",
    "8.11": "Entreprise : annuler un transfert.",
    "8.12": "Tenter d'annuler un transfert déjà exécuté → refusé.",
    "8.13": "Entreprise : vérifier la notification.",
    # Phase 9
    "9.1": "Entreprise → Opérations > Sous-traitance > Onboarding : créer le compte sous-traitant.",
    "9.2": "Sous-traitance : soumettre une demande.",
    "9.3": "DGTCP → Sous-traitance : la file s'affiche.",
    "9.4": "DGTCP : autoriser la sous-traitance.",
    "9.5": "DGTCP : refuser la sous-traitance.",
    "9.6": "Entreprise : suspendre le sous-traitant.",
    "9.7": "Entreprise : réactiver le sous-traitant.",
    "9.8": "Entreprise : révoquer (action irréversible).",
    "9.9": "Se connecter en sous-traitant : seul le certificat partagé est visible.",
    "9.10": "Sous-traitant : créer une utilisation dans la limite du solde alloué.",
    "9.11": "Entreprise : consulter la liste des sous-traitants.",
    # Phase 10
    "10.1": "DGTCP → Opérations > Clôture : la file des certificats éligibles s'affiche avec les motifs.",
    "10.2": "Clôture : consulter les certificats immédiatement clôturables.",
    "10.3": "DGTCP : proposer une clôture.",
    "10.4": "Détail clôture : joindre les documents.",
    "10.5": "Président → Clôture : la proposition s'affiche.",
    "10.6": "Président : valider la clôture.",
    "10.7": "Président : rejeter la clôture.",
    "10.8": "DGTCP : finaliser → certificat clôturé.",
    "10.9": "Tenter une utilisation sur le certificat clôturé → refusé.",
    "10.10": "Entreprise : vérifier la notification + l'e-mail.",
    # Phase 11
    "11.1": "Sur un dossier, téléverser un document (version 1).",
    "11.2": "Remplacer ce document → version 2 créée.",
    "11.3": "Consulter le dossier GED : v1 et v2 présentes (v1 inactive).",
    "11.4": "Admin → GED > Dossiers : les dossiers s'affichent.",
    "11.5": "GED > Dossiers : ouvrir un dossier pour voir l'arborescence et les versions.",
    "11.6": "GED > Configuration : créer un type de document + une exigence.",
    # Phase 12
    "12.1.1": "Cliquer sur la cloche (haut à droite) : la liste des notifications s'affiche.",
    "12.1.2": "Vérifier le badge compteur de non-lus sur la cloche.",
    "12.1.3": "Cliquer sur une notification pour la marquer comme lue.",
    "12.1.4": "Bouton 'Tout marquer comme lu'.",
    "12.2.1": "Déclencher une soumission de correction, puis vérifier les boîtes mail (Président + DGD + Commission).",
    "12.2.2": "Déclencher une adoption, puis vérifier les boîtes mail (AC + Entreprise).",
    "12.2.3": "Déclencher l'ouverture d'un certificat, puis vérifier la boîte mail (Entreprise).",
    "12.2.4": "Déclencher une liquidation, puis vérifier la boîte mail (Entreprise).",
    "12.2.5": "Déclencher une décision de transfert, puis vérifier la boîte mail (Entreprise).",
    "12.2.6": "Déclencher une clôture, puis vérifier la boîte mail (Entreprise).",
    "12.2.7": "Approuver un reset de mot de passe, puis vérifier la boîte mail du demandeur.",
    "12.3.1": "Commission Relais → écran Relais : la liste des entreprises s'affiche.",
    "12.3.2": "Sélectionner une entreprise pour agir en son nom (impersonation).",
    "12.3.3": "Effectuer une action (créer une utilisation) en tant qu'entreprise.",
    "12.3.4": "Bouton 'Quitter la session' (release) → retour au compte Commission Relais.",
    "12.3.5": "Sélectionner une autorité contractante pour agir en son nom.",
    # Phase 13
    "13.1": "Président → Reporting : la synthèse nationale s'affiche.",
    "13.2": "DGTCP → Reporting : filtrer par entreprise.",
    "13.3": "AC → Reporting : données limitées à son périmètre.",
    "13.4": "Entreprise → Reporting : uniquement ses propres dossiers.",
    "13.5": "Reporting : consulter la série temporelle des demandes.",
    "13.6": "Reporting : appliquer un filtre de période.",
    "13.7": "Admin → Paramétrage > Audit : le journal d'audit s'affiche.",
    "13.8": "Audit : ouvrir une entrée pour consulter le snapshot de l'objet.",
    # Phase 14
    "14.1": "Saisir un montant égal au solde restant → accepté, solde à 0.",
    "14.2": "Soumettre deux fois le même brouillon → la 2e tentative est refusée.",
    "14.3": "Tenter de viser hors ordre (DGTCP avant DGI) → refusé.",
    "14.4": "Tenter une utilisation sur un certificat non ouvert → refusé.",
    "14.5": "Téléverser un fichier vide → refusé.",
    "14.6": "Avec un compte d'une autre entreprise, tenter d'accéder aux ressources d'une autre → refusé.",
    "14.7": "Reporting : appliquer un filtre avec une date très ancienne → pas d'erreur.",
    "14.8": "Couper le serveur mail puis déclencher une action : l'opération métier réussit malgré tout.",
    "14.9": "Ouvrir deux rejets temporaires : les deux doivent être résolus avant de continuer.",
    "14.10": "Tenter une liquidation supérieure au solde → message d'erreur clair.",
    "14.11": "Après révocation d'un sous-traitant, tenter une utilisation → accès refusé.",
    "14.12": "Envoyer un message sur une demande d'explication fermée → refusé.",
}

# ── Construction Excel ─────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Plan de Test"

# ── Validation liste déroulante statut ────────────────────────────────────────
dv = DataValidation(
    type="list",
    formula1='"A TESTER,EN COURS,PASSE,ECHOUE,BLOQUE,N/A"',
    allow_blank=True,
    showDropDown=False,
    showErrorMessage=True,
    error="Valeur invalide",
    errorTitle="Statut"
)
ws.add_data_validation(dv)

# ── En-têtes colonnes ─────────────────────────────────────────────────────────
headers = ["#", "Nom du test", "Comment tester (UI)", "Action / Endpoint", "Résultat attendu", "Statut", "Commentaire", "Testeur", "Date"]
col_widths = [10, 28, 52, 46, 40, 14, 32, 14, 13]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font()
    cell.fill = make_fill(C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"

# ── Remplissage lignes ─────────────────────────────────────────────────────────
row_num = 2
test_count = 0
phase_count = 0

for entry in ROWS:
    kind = entry[0]

    if kind == "PHASE":
        phase_count += 1
        cell = ws.cell(row=row_num, column=1, value=entry[1])
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
        cell.font = Font(name="Calibri", bold=True, size=12, color=C_WHITE)
        cell.fill = make_fill(C_PHASE_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = thin_border()
        ws.row_dimensions[row_num].height = 22

    elif kind == "SECTION":
        cell = ws.cell(row=row_num, column=1, value=entry[1])
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
        cell.font = Font(name="Calibri", bold=True, size=10, color="1F4E79")
        cell.fill = make_fill(C_SECTION_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        cell.border = thin_border()
        ws.row_dimensions[row_num].height = 18

    elif kind == "TEST":
        test_count += 1
        _, num, nom, action, attendu = entry
        description = DESCRIPTIONS.get(num, "")
        alt = (test_count % 2 == 0)
        bg = C_GREY_ROW if alt else C_WHITE

        values = [num, nom, description, action, attendu, "A TESTER", "", "", ""]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = cell_font()
            cell.fill = make_fill(bg)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="top")
                cell.font = cell_font(bold=True, color="1F4E79")
            if col_idx == 3:  # Comment tester (UI)
                cell.font = cell_font(color="375623")
            if col_idx == 6:  # Statut
                cell.alignment = Alignment(horizontal="center", vertical="top")
                dv.add(cell)
        ws.row_dimensions[row_num].height = 48

    row_num += 1

# ── Mise en forme conditionnelle (couleur selon statut) ───────────────────────
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

last_row = row_num - 1
stat_col = "F"

rules = [
    ("PASSE",   C_LIGHT_GREEN, C_GREEN),
    ("ECHOUE",  C_LIGHT_RED,   C_RED),
    ("EN COURS",C_LIGHT_YELL,  "BF8F00"),
    ("BLOQUE",  "FFE0CC",      "C55A11"),
    ("N/A",     "EFEFEF",      "7F7F7F"),
]
for statut, bg, font_col in rules:
    fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    font = Font(color=font_col, bold=True, name="Calibri", size=10)
    ws.conditional_formatting.add(
        f"{stat_col}2:{stat_col}{last_row}",
        CellIsRule(operator="equal", formula=[f'"{statut}"'], fill=fill, font=font)
    )

# ── Onglet Résumé ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Résumé")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 16

# Titre
title_cell = ws2.cell(row=1, column=1, value="Tableau de bord — SGCI Test Post-Livraison")
ws2.merge_cells("A1:C1")
title_cell.font = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
title_cell.fill = make_fill(C_HEADER_BG)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

# Sous-titre
for col, val in [(1, "Phase"), (2, "Tests"), (3, "% Terminé (formule)")]:
    cell = ws2.cell(row=2, column=col, value=val)
    cell.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
    cell.fill = make_fill(C_PHASE_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
ws2.row_dimensions[2].height = 20

phases = [
    ("Phase 1 — Infrastructure & Sécurité", 10),
    ("Phase 2 — Référentiels & Administration", 17),
    ("Phase 3 — Convention & Marché", 16),
    ("Phase 4 — Demande de Correction", 23),
    ("Phase 5 — Mise en place Certificat", 15),
    ("Phase 6 — Utilisation Douane", 19),
    ("Phase 7 — Utilisation TVA", 12),
    ("Phase 8 — Transfert de Crédit", 13),
    ("Phase 9 — Sous-traitance", 11),
    ("Phase 10 — Clôture du Certificat", 10),
    ("Phase 11 — GED & Documents", 6),
    ("Phase 12 — Notifications & Commission Relais", 16),
    ("Phase 13 — Reporting & Audit", 8),
    ("Phase 14 — Régression & Cas limites", 12),
]

for i, (phase_name, nb_tests) in enumerate(phases, start=3):
    alt = (i % 2 == 0)
    bg = C_GREY_ROW if alt else C_WHITE
    c1 = ws2.cell(row=i, column=1, value=phase_name)
    c2 = ws2.cell(row=i, column=2, value=nb_tests)
    c3 = ws2.cell(row=i, column=3, value="à calculer manuellement")
    for c in [c1, c2, c3]:
        c.font = cell_font(size=10)
        c.fill = make_fill(bg)
        c.border = thin_border()
        c.alignment = Alignment(vertical="center")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c3.font = cell_font(size=10, color="7F7F7F")
    ws2.row_dimensions[i].height = 18

# Total
total_row = len(phases) + 3
ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11)
ws2.cell(row=total_row, column=2, value=sum(p[1] for p in phases)).font = Font(name="Calibri", bold=True, size=11)
ws2.cell(row=total_row, column=1).fill = make_fill(C_SECTION_BG)
ws2.cell(row=total_row, column=2).fill = make_fill(C_SECTION_BG)
ws2.cell(row=total_row, column=3).fill = make_fill(C_SECTION_BG)
for col in range(1, 4):
    ws2.cell(row=total_row, column=col).border = thin_border()
    ws2.cell(row=total_row, column=col).alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[total_row].height = 22

# Légende
legend_start = total_row + 2
ws2.cell(row=legend_start, column=1, value="Légende Statuts").font = Font(name="Calibri", bold=True, size=11)
legends = [
    ("A TESTER",  C_WHITE,      "000000", "Non démarré"),
    ("EN COURS",  C_LIGHT_YELL, "BF8F00", "Test en cours"),
    ("PASSE",     C_LIGHT_GREEN,C_GREEN,  "Test réussi"),
    ("ECHOUE",    C_LIGHT_RED,  C_RED,    "Test échoué — bug à corriger"),
    ("BLOQUE",    "FFE0CC",     "C55A11", "Bloqué par prérequis"),
    ("N/A",       "EFEFEF",     "7F7F7F", "Non applicable"),
]
for j, (label, bg, fc, desc) in enumerate(legends, start=legend_start + 1):
    c = ws2.cell(row=j, column=1, value=label)
    c.font = Font(name="Calibri", bold=True, size=10, color=fc)
    c.fill = make_fill(bg)
    c.border = thin_border()
    c.alignment = Alignment(horizontal="center", vertical="center")
    d = ws2.cell(row=j, column=2, value=desc)
    d.font = cell_font()
    ws2.merge_cells(start_row=j, start_column=2, end_row=j, end_column=3)
    d.alignment = Alignment(vertical="center")
    ws2.row_dimensions[j].height = 18

# ── Sauvegarde ─────────────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(OUT), exist_ok=True)
try:
    wb.save(OUT)
    target = OUT
except PermissionError:
    import time
    base, ext = os.path.splitext(OUT)
    target = f"{base}_{time.strftime('%Y%m%d_%H%M%S')}{ext}"
    wb.save(target)
    print("[WARN] Fichier principal verrouille (ouvert dans Excel). Copie enregistree a la place.")
print(f"[OK] Fichier genere : {target}")
print(f"     {test_count} tests, {phase_count} phases")
